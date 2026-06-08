"""
Formatting Script: Generate XLS and PDF from Verified Entries
For Haroon & Sons Coconut Brokerage Bill Automation
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

class BillFormatter:
    def __init__(self, client_name: str, brokerage_rate: int, bill_period_start: str, bill_period_end: str):
        """
        Initialize formatter with bill details.
        
        Args:
            client_name: e.g., "M/S LALCHAND RAMCHAND, VASHI"
            brokerage_rate: 5 or 10 (rupees per katta)
            bill_period_start: DD-MM-YYYY
            bill_period_end: DD-MM-YYYY
        """
        self.client_name = client_name
        self.brokerage_rate = brokerage_rate
        self.bill_period_start = bill_period_start
        self.bill_period_end = bill_period_end
        self.company_name = "HAROON & SONS"
        self.company_subtitle = "COPRA & OIL BROKERS"
        self.company_address = "OFFICE: 371/73, Narshi Natha Street, 2/A. Faize-E-Edroos Building, Mumbai - 400 009"
    
    def prepare_entries(self, validated_entries: List[Dict]) -> List[Dict]:
        """
        Prepare entries for output: calculate brokerage, sort by date.
        """
        prepared = []
        
        for entry in validated_entries:
            original = entry['original_entry']
            
            try:
                katta = int(original['katta'])
                brokerage = katta * self.brokerage_rate
            except:
                brokerage = 0
            
            prepared.append({
                'date': original['date'],
                'katta': original['katta'],
                'rate': original['rate'],
                'party': original['party'],
                'bill_details': original.get('bill_details', ''),
                'brokerage': brokerage
            })
        
        # Sort by date
        prepared.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"))
        
        return prepared
    
    def generate_xls(self, prepared_entries: List[Dict], output_path: str) -> str:
        """
        Generate Excel file with proper formatting.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Bill"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Date
        ws.column_dimensions['B'].width = 12  # Katta
        ws.column_dimensions['C'].width = 12  # Rate
        ws.column_dimensions['D'].width = 25  # Party
        ws.column_dimensions['E'].width = 15  # Details
        ws.column_dimensions['F'].width = 15  # Brokerage
        
        # Define styles
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add company header
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_name
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_subtitle
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = self.company_address
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add client info
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = self.client_name
        cell.font = Font(bold=True, size=11)
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f"BILL PERIOD FROM: {self.bill_period_start} UPTO {self.bill_period_end}"
        cell.font = Font(size=10)
        
        # Add column headers
        row += 2
        headers = ['DATE', 'KATTA', 'RATE', 'PARTY', 'DETAILS', 'BROKERAGE']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add data rows
        row += 1
        total_brokerage = 0
        for entry in prepared_entries:
            ws.cell(row=row, column=1).value = entry['date']
            ws.cell(row=row, column=2).value = int(entry['katta'])
            ws.cell(row=row, column=3).value = float(entry['rate'])
            ws.cell(row=row, column=4).value = entry['party']
            ws.cell(row=row, column=5).value = entry['bill_details']
            ws.cell(row=row, column=6).value = entry['brokerage']
            
            total_brokerage += entry['brokerage']
            
            # Style data row
            for col_num in range(1, 7):
                cell = ws.cell(row=row, column=col_num)
                cell.border = border
                if col_num in [2, 3, 6]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
            
            row += 1
        
        # Add total row
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "TOTAL"
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        
        cell = ws[f'F{row}']
        cell.value = total_brokerage
        cell.font = Font(bold=True, size=11)
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        
        wb.save(output_path)
        return output_path
    


# Test the formatter
if __name__ == "__main__":
    sample_prepared = [
        {
            'date': '07-04-2025',
            'katta': '400',
            'rate': '218',
            'party': 'M/S MUSKHAN ENTERPRISES',
            'bill_details': 'Bill 6',
            'brokerage': 2000
        },
        {
            'date': '12-04-2025',
            'katta': '400',
            'rate': '202',
            'party': 'M/S SHAAN ENTERPRISES',
            'bill_details': 'Bill 124',
            'brokerage': 2000
        }
    ]
    
    formatter = BillFormatter(
        client_name="M/S LALCHAND RAMCHAND, VASHI",
        brokerage_rate=5,
        bill_period_start="01-04-2025",
        bill_period_end="31-03-2026"
    )
    
    # Generate XLS
    xls_path = formatter.generate_xls(sample_prepared, '/tmp/test_bill.xlsx')
    print(f"XLS generated: {xls_path}")
    
    # Generate PDF
    pdf_path = formatter.generate_pdf(sample_prepared, '/tmp/test_bill.pdf')
    print(f"PDF generated: {pdf_path}")
